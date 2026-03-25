#!/usr/bin/env python
# coding: utf-8

# #!/usr/bin/env python3
# # -*- coding: utf-8 -*-
# 
# """
# US dynamic IO model (Step A-C)
# - Input files expected in current directory:
#     * Use table.xlsx
#     * Import Matrix.xlsx
# 
# Implemented scope:
# - Step A: build US baseline from BEA Use table and Import Matrix
# - Step B: no-shock simulation for steady-state sanity check
# - Step C: linear production function + single-sector supply shock
# 
# Important modeling choices in this script:
# 1) We use the BEA "After Redefinitions - Summary" Use table.
# 2) We keep only the 71 matched row/column industries and drop:
#       - Used
#       - Other
#    because they are extra commodity rows rather than ordinary industries.
# 3) We build domestic intermediate use as:
#       Z_domestic = Use_intermediate - Import_intermediate
#    so imports do not create bottlenecks.
# 4) Because the Use table is commodity-by-industry, the internally consistent
#    baseline output vector for demand balancing is the row-side commodity output,
#    not the column-side total industry output row. Labor compensation is still
#    taken from the industry columns and matched by BEA code.
# 5) Daily model: annual BEA flows (million USD/year) are converted to
#    million USD/day by dividing by DAYS_PER_YEAR.
# 6) This is intentionally a minimal Step A-C build:
#    - no household income feedback
#    - no endogenous labor hiring/firing
#    - no demand shock process
#    - no PBL / critical input logic yet
# 7) Hooks for future import constraints are preserved in the data structures.
# """

# In[1]:


from __future__ import annotations

import math
import json
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

import pickle

import warnings
warnings.simplefilter('ignore')

from IPython.display import display


# In[2]:


def load_io_core(path: str) -> Dict:
    with open(path, "rb") as f:
        return pickle.load(f)


# In[3]:


# =========================
# User settings
# =========================

USE_FILE = "Use table.xlsx"
IMPORT_FILE = "Import Matrix.xlsx"

IO_CORE_FILE = "io_core.pkl"

DAYS_PER_YEAR = 365.0

# Uniform inventory target in "days of baseline intermediate use"
INVENTORY_DAYS = 5.0

# Inventory gap adjustment parameter
TAU = 10.0

# Simulation length
T_NO_SHOCK = 60
T_SHOCK = 120

# Shock scenario
SHOCK_INDUSTRY_CODE = "325"
SHOCK_START = 10
SHOCK_END = 40
SHOCK_SIZE = 0.30

# Production function mode
# "linear" / "pseudo_pbl" / "pbl"
#PRODUCTION_MODE = "linear"
#PRODUCTION_MODE = "pseudo_pbl"
PRODUCTION_MODE = "pbl"

# Comparison modes to run together
COMPARE_MODES = ["linear", "pseudo_pbl", "pbl"]

# Pseudo-PBL classification parameters
CRITICAL_CUM_SHARE = 0.05
IMPORTANT_CUM_SHARE = 0.80
MIN_CRITICAL_INPUTS = 1

OUT_DIR = Path("output_stepD")
OUT_DIR.mkdir(exist_ok=True)

# Input-network construction mode
# "mechanical" / "domain_knowledge" / "hybrid"
NETWORK_MODE = "domain_knowledge"


# In[4]:


# =========================
# Helpers
# =========================

def to_float(x) -> float:
    """Convert BEA cell contents to float; treat --- / blank as 0."""
    if x is None:
        return 0.0
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---"}:
            return 0.0
        x = x.replace(",", "")
        return float(x)
    return float(x)


def safe_divide(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    out = np.zeros_like(a, dtype=float)
    np.divide(a, b, out=out, where=np.abs(b) > 1e-15)
    return out


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)

def build_pseudo_pbl_masks(
    A: np.ndarray,
    critical_cum_share: float = CRITICAL_CUM_SHARE,
    important_cum_share: float = IMPORTANT_CUM_SHARE,
    min_critical_inputs: int = MIN_CRITICAL_INPUTS,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:

    N = A.shape[0]
    critical = np.zeros_like(A, dtype=bool)
    important = np.zeros_like(A, dtype=bool)
    noncritical = np.zeros_like(A, dtype=bool)

    for i in range(N):
        col = A[:, i]
        pos_idx = np.where(col > 1e-15)[0]
        if len(pos_idx) == 0:
            continue

        shares = col[pos_idx]
        order = np.argsort(-shares)
        sorted_idx = pos_idx[order]
        sorted_shares = shares[order]

        cum = np.cumsum(sorted_shares) / sorted_shares.sum()

        # 修正：thresholdを超えるまで含める
        crit_k = np.searchsorted(cum, critical_cum_share, side="left") + 1
        crit_k = max(min_critical_inputs, crit_k)
        crit_k = min(crit_k, len(sorted_idx))

        imp_k = np.searchsorted(cum, important_cum_share, side="left") + 1
        imp_k = max(imp_k, crit_k)
        imp_k = min(imp_k, len(sorted_idx))

        crit_idx = sorted_idx[:crit_k]
        imp_idx = sorted_idx[crit_k:imp_k]
        non_idx = sorted_idx[imp_k:]

        critical[crit_idx, i] = True
        important[imp_idx, i] = True
        noncritical[non_idx, i] = True

    return critical, important, noncritical


def compute_input_constraint(
    S: np.ndarray,
    A: np.ndarray,
    mode: str,
    critical_mask: Optional[np.ndarray] = None,
    important_mask: Optional[np.ndarray] = None,
) -> np.ndarray:
    """
    Compute x_inp under either:
    - linear
    - pseudo_pbl
    """
    N = A.shape[1]

    if mode == "linear":
        colsum_A = A.sum(axis=0)
        x_inp = np.full(N, np.inf, dtype=float)
        positive_recipe = colsum_A > 1e-15
        x_inp[positive_recipe] = S[:, positive_recipe].sum(axis=0) / colsum_A[positive_recipe]
        return x_inp

    if mode == "pseudo_pbl":
        if critical_mask is None or important_mask is None:
            raise ValueError("pseudo_pbl mode requires critical_mask and important_mask")

        x_inp = np.full(N, np.inf, dtype=float)

        for i in range(N):
            caps = []

            crit_idx = np.where(critical_mask[:, i] & (A[:, i] > 1e-15))[0]
            if len(crit_idx) > 0:
                caps.append(np.min(S[crit_idx, i] / A[crit_idx, i]))

            imp_idx = np.where(important_mask[:, i] & (A[:, i] > 1e-15))[0]
            if len(imp_idx) > 0:
                caps.append(S[imp_idx, i].sum() / A[imp_idx, i].sum())

            # fallback if a sector has recipe but no classified critical/important inputs
            pos_idx = np.where(A[:, i] > 1e-15)[0]
            if len(caps) == 0 and len(pos_idx) > 0:
                caps.append(S[pos_idx, i].sum() / A[pos_idx, i].sum())

            x_inp[i] = min(caps) if len(caps) > 0 else np.inf

        return x_inp

    raise ValueError(f"Unknown production mode: {mode}")
    
def compute_input_constraint_pseudo_pbl(
    S: np.ndarray,
    A: np.ndarray,
    critical_mask: np.ndarray,
    important_mask: np.ndarray,
) -> np.ndarray:

    N = A.shape[0]
    x_inp = np.full(N, np.inf)

    for i in range(N):
        caps = []

        # --- critical: 完全Leontief ---
        crit_idx = np.where(critical_mask[:, i] & (A[:, i] > 1e-15))[0]
        for j in crit_idx:
            caps.append(S[j, i] / A[j, i])

        # --- important: ここを変更（合算→個別制約） ---
        imp_idx = np.where(important_mask[:, i] & (A[:, i] > 1e-15))[0]
        for j in imp_idx:
            # 少し緩めたい場合は係数をかける
            caps.append(0.7 * S[j, i] / A[j, i])

        # fallback（投入ゼロ列対策）
        if len(caps) == 0:
            colsum = A[:, i].sum()
            if colsum > 1e-15:
                caps.append(S[:, i].sum() / colsum)

        x_inp[i] = min(caps) if caps else np.inf

    return x_inp

def compute_input_constraint_pbl(
    S: np.ndarray,
    A: np.ndarray,
    critical_mask: np.ndarray,
) -> np.ndarray:
    """
    True PBL:
    output is constrained only by critical inputs via pure min operator.
    """
    N = A.shape[1]
    x_inp = np.full(N, np.inf, dtype=float)

    for i in range(N):
        crit_idx = np.where(critical_mask[:, i] & (A[:, i] > 1e-15))[0]

        if len(crit_idx) > 0:
            x_inp[i] = np.min(S[crit_idx, i] / A[crit_idx, i])
        else:
            # fallback if no critical inputs are assigned
            pos_idx = np.where(A[:, i] > 1e-15)[0]
            if len(pos_idx) > 0:
                x_inp[i] = np.min(S[pos_idx, i] / A[pos_idx, i])
            else:
                x_inp[i] = np.inf

    return x_inp


# In[5]:


def build_domain_knowledge_masks(
    codes: List[str],
    names: List[str],
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Sparse domain-knowledge network for BEA 71-sector summary codes.

    Principle:
    - critical: only hard physical bottlenecks
      (goods/materials, logistics, utilities/core digital infra)
    - important: used sparingly, only where short-run disruption is plausible
    - finance / real estate / most business services: noncritical by default
    """
    N = len(codes)
    code_to_idx = {c: i for i, c in enumerate(codes)}

    critical = np.zeros((N, N), dtype=bool)
    important = np.zeros((N, N), dtype=bool)

    manufacturing = {
        "311FT", "313TT", "315AL", "321", "322", "323", "324", "325", "326", "327",
        "331", "332", "333", "334", "335", "3361MV", "3364OT", "337", "339"
    }
    goods = manufacturing | {"111CA", "113FF", "211", "212", "213"}
    transport = {"42", "481", "482", "483", "484", "485", "486", "48A", "493"}
    retail = {"441", "445", "452", "4A0"}
    health = {"621", "622", "623", "624"}
    digital_core = {"513", "514"}
    construction_users = {"23"} | manufacturing

    def set_crit(input_code: str, users: set):
        if input_code not in code_to_idx:
            return
        j = code_to_idx[input_code]
        for u in users:
            if u in code_to_idx:
                i = code_to_idx[u]
                critical[j, i] = True
                important[j, i] = False

    def set_imp(input_code: str, users: set):
        if input_code not in code_to_idx:
            return
        j = code_to_idx[input_code]
        for u in users:
            if u in code_to_idx:
                i = code_to_idx[u]
                if not critical[j, i]:
                    important[j, i] = True

    # -------------------------------------------------
    # 1. Utilities and core digital infrastructure
    # -------------------------------------------------
    # Utilities are near-universal short-run bottlenecks
    set_crit("22", set(codes) - {"22"})

    # Telecom / data infra: critical only for sectors that truly cannot function without it
    set_crit("513", {
        "42", "481", "482", "483", "484", "485", "486", "48A", "493",
        "52FR", "523", "524", "525", "511", "514", "5415",
        "GFGD", "GFGN", "GFE", "GSLG", "GSLE"
    })
    set_crit("514", {
        "42", "481", "482", "483", "484", "485", "486", "48A", "493",
        "52FR", "523", "524", "525", "511", "513", "5415",
        "GFGD", "GFGN", "GFE", "GSLG", "GSLE"
    })

    # -------------------------------------------------
    # 2. Logistics bottlenecks
    # -------------------------------------------------
    set_crit("42", goods | {"23"} | transport | retail)
    set_crit("484", goods | {"23"} | transport | retail)
    set_imp("493", manufacturing | {"23", "42", "441", "445", "452", "4A0"})
    set_imp("48A", {"42", "481", "482", "483", "484", "486", "493"})

    # -------------------------------------------------
    # 3. Core physical inputs
    # -------------------------------------------------
    # Energy / refined petroleum
    set_crit("324", goods | {"23"} | transport | retail)

    # Chemicals
    set_crit("325", goods | {"23"} | health)

    # Basic materials / industrial inputs
    set_crit("331", manufacturing | {"23"})
    set_crit("332", manufacturing | {"23"})
    set_crit("333", manufacturing | {"23"})
    set_crit("334", {"333", "334", "335", "3361MV", "3364OT", "339"})
    set_crit("335", manufacturing | {"23"})

    # Inputs that matter but are not universal bottlenecks
    set_imp("326", manufacturing | {"23"})
    set_imp("327", {"23", "321", "324", "325", "331", "332", "333", "335"})
    set_imp("321", {"23", "321", "322", "323", "337"})
    set_imp("322", {"323", "445", "452", "4A0"})

    # Agriculture / food chain
    set_crit("111CA", {"311FT"})
    set_crit("311FT", {"445", "722"})
    set_imp("113FF", {"321", "337", "322"})

    # Transport equipment
    set_crit("3361MV", {"441", "484", "485"})
    set_imp("3364OT", {"481", "483", "GFGD", "GFE"})

    # Extractives as upstream important, not broadly critical
    set_imp("211", {"324", "325", "486"})
    set_imp("212", {"327", "331", "332", "333"})
    set_imp("213", {"211", "212"})

    # -------------------------------------------------
    # 4. Construction / health
    # -------------------------------------------------
    set_crit("23", {"531", "HS", "GFGD", "GFGN", "GSLG", "GSLE"})
    set_crit("621", health)
    set_crit("622", {"623", "624"})
    set_imp("623", {"624"})

    # -------------------------------------------------
    # 5. Everything else is noncritical by default
    # -------------------------------------------------
    noncritical = ~(critical | important)
    return critical, important, noncritical


# In[6]:


def build_input_network(
    mode: str,
    A: np.ndarray,
    codes: List[str],
    names: List[str],
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Unified network builder:
    - mechanical
    - domain_knowledge
    - hybrid
    """
    mech_crit, mech_imp, mech_non = build_pseudo_pbl_masks(
        A=A,
        critical_cum_share=CRITICAL_CUM_SHARE,
        important_cum_share=IMPORTANT_CUM_SHARE,
        min_critical_inputs=MIN_CRITICAL_INPUTS,
    )

    if mode == "mechanical":
        return mech_crit, mech_imp, mech_non

    dom_crit, dom_imp, dom_non = build_domain_knowledge_masks(codes, names)

    if mode == "domain_knowledge":
        return dom_crit, dom_imp, dom_non

    if mode == "hybrid":
        # Domain knowledge takes priority for critical links.
        # Mechanical fills the rest.
        crit = dom_crit | (mech_crit & ~dom_imp & ~dom_crit)
        imp = dom_imp | (mech_imp & ~crit)
        non = ~(crit | imp)
        return crit, imp, non

    raise ValueError(f"Unknown NETWORK_MODE: {mode}")


# In[7]:


# =========================
# Data structures
# =========================

@dataclass
class BEABaseline:
    codes: List[str]
    names: List[str]

    # Annual flows (million USD / year)
    U_total_annual: np.ndarray
    M_import_annual: np.ndarray
    c0_annual: np.ndarray
    f0_annual: np.ndarray
    x0_annual: np.ndarray
    l0_annual: np.ndarray

    # Daily flows (million USD / day)
    U_total_daily: np.ndarray
    M_import_daily: np.ndarray
    c0_daily: np.ndarray
    f0_daily: np.ndarray
    x0_daily: np.ndarray
    l0_daily: np.ndarray

    # Technical coefficients
    A_total: np.ndarray

    # Inventories
    n_days: np.ndarray
    S0_total: np.ndarray

    # Pseudo-PBL masks
    critical_mask: np.ndarray
    important_mask: np.ndarray
    noncritical_mask: np.ndarray

    # Future extension only
    import_share_matrix: np.ndarray
    import_share_col: np.ndarray

    # Diagnostics only
    negative_domestic_entries: List[Dict]
    negative_domestic_count: int


@dataclass
class SimulationResult:
    t: np.ndarray
    x: np.ndarray
    d: np.ndarray
    x_cap: np.ndarray
    x_inp: np.ndarray
    c_realized: np.ndarray
    f_realized: np.ndarray
    shock_multiplier: np.ndarray
    aggregate_output: np.ndarray
    aggregate_demand: np.ndarray
    output_ratio_to_baseline: np.ndarray
    sector_output_ratio: np.ndarray


# In[8]:


# =========================
# Excel parsing
# =========================

def read_use_table(path: str) -> Dict:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb["Table"]

    # Industry columns in summary table: C:BU = 71 industries
    industry_col_start = 3
    industry_col_end = 73

    # Main commodity rows:
    # row 8 to row 78 are the 71 matched industries/commodities
    # row 79 = Used, row 80 = Other -> exclude
    row_start = 8
    row_end = 78

    codes = [ws.cell(row_start + i, 1).value for i in range(row_end - row_start + 1)]
    names = [ws.cell(row_start + i, 2).value for i in range(row_end - row_start + 1)]
    col_codes = [ws.cell(6, c).value for c in range(industry_col_start, industry_col_end + 1)]
    col_names = [ws.cell(7, c).value for c in range(industry_col_start, industry_col_end + 1)]

    if codes != col_codes:
        raise ValueError("Row codes and column codes do not align for the 71-industry block.")

    # Intermediate use block (rows: commodities, cols: industries)
    U_total = np.array(
        [
            [to_float(ws.cell(r, c).value) for c in range(industry_col_start, industry_col_end + 1)]
            for r in range(row_start, row_end + 1)
        ],
        dtype=float,
    )

    # Final demand
    # 75: Personal consumption expenditures
    # 95: Total Final Uses (GDP)
    col_pce = 75
    col_total_final_uses = 95

    c0 = np.array(
        [to_float(ws.cell(r, col_pce).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )

    total_final_uses = np.array(
        [to_float(ws.cell(r, col_total_final_uses).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )
    
    # Other final demand excluding household consumption
    f0 = total_final_uses - c0
    f0[np.abs(f0) < 1e-12] = 0.0

    # Row-side total commodity output
    # In this sheet col 96 contains "Total Commodity Output"
    col_total_commodity_output = 96
    x0 = np.array(
        [to_float(ws.cell(r, col_total_commodity_output).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )

    # Industry-side compensation of employees row
    row_compensation = 82
    l0 = np.array(
        [to_float(ws.cell(row_compensation, c).value) for c in range(industry_col_start, industry_col_end + 1)],
        dtype=float,
    )

    # Industry-side output row (not used for simulation baseline, but useful for diagnostics)
    row_total_industry_output = 86
    x_industry = np.array(
        [to_float(ws.cell(row_total_industry_output, c).value) for c in range(industry_col_start, industry_col_end + 1)],
        dtype=float,
    )

    return {
        "codes": codes,
        "names": names,
        "U_total": U_total,
        "c0": c0,
        "f0": f0,
        "x0_commodity": x0,
        "l0": l0,
        "x0_industry": x_industry,
        "col_names": col_names,
    }

def read_import_matrix(path: str, expected_codes: List[str]) -> np.ndarray:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb["Table"]

    industry_col_start = 3
    industry_col_end = 73
    row_start = 8
    row_end = 78

    row_codes = [ws.cell(row_start + i, 1).value for i in range(row_end - row_start + 1)]
    col_codes = [ws.cell(6, c).value for c in range(industry_col_start, industry_col_end + 1)]

    if row_codes != expected_codes or col_codes != expected_codes:
        raise ValueError("Import matrix codes do not align with Use table 71-industry block.")

    M_import = np.array(
        [
            [to_float(ws.cell(r, c).value) for c in range(industry_col_start, industry_col_end + 1)]
            for r in range(row_start, row_end + 1)
        ],
        dtype=float,
    )

    return M_import


# In[9]:


# =========================
# Build baseline
# =========================

def build_baseline_from_core(io_core_file: str) -> BEABaseline:
    core = load_io_core(io_core_file)

    U_total = core["U_total_annual"]
    M_import = core["M_import_annual"]

    # Diagnostics only for future import extension
    Z_dom_raw = U_total - M_import
    Z_dom_raw[np.abs(Z_dom_raw) < 1e-12] = 0.0

    neg_mask = Z_dom_raw < 0
    neg_count = int(neg_mask.sum())
    neg_entries = []

    if neg_count > 0:
        neg_idx = np.argwhere(neg_mask)
        for r, c in neg_idx:
            neg_entries.append({
                "row_code": core["codes"][r],
                "row_name": core["names"][r],
                "col_code": core["codes"][c],
                "col_name": core["col_names"][c],
                "use_value": float(U_total[r, c]),
                "import_value": float(M_import[r, c]),
                "domestic_raw": float(Z_dom_raw[r, c]),
            })

    # Rebuild only parameter-dependent objects
    n_days = np.full(len(core["codes"]), INVENTORY_DAYS, dtype=float)
    S0_total = core["U_total_daily"] * n_days.reshape(1, -1)

    critical_mask, important_mask, noncritical_mask = build_input_network(
        mode=NETWORK_MODE,
        A=core["A_total"],
        codes=core["codes"],
        names=core["names"],
    )

    baseline = BEABaseline(
        codes=core["codes"],
        names=core["names"],
        U_total_annual=core["U_total_annual"],
        M_import_annual=core["M_import_annual"],
        c0_annual=core["c0_annual"],
        f0_annual=core["f0_annual"],
        x0_annual=core["x0_annual"],
        l0_annual=core["l0_annual"],
        U_total_daily=core["U_total_daily"],
        M_import_daily=core["M_import_daily"],
        c0_daily=core["c0_daily"],
        f0_daily=core["f0_daily"],
        x0_daily=core["x0_daily"],
        l0_daily=core["l0_daily"],
        A_total=core["A_total"],
        n_days=n_days,
        S0_total=S0_total,
        critical_mask=critical_mask,
        important_mask=important_mask,
        noncritical_mask=noncritical_mask,
        import_share_matrix=core["import_share_matrix"],
        import_share_col=core["import_share_col"],
        negative_domestic_entries=neg_entries,
        negative_domestic_count=neg_count,
    )
    return baseline


# In[10]:


# =========================
# Simulation engine (Step B-C)
# =========================


def build_supply_shock_path(
    codes: List[str],
    T: int,
    shock_code: Optional[str] = None,
    shock_start: int = 0,
    shock_end: int = -1,
    shock_size: float = 0.0,
) -> np.ndarray:
    """
    Returns a multiplicative labor/capacity factor matrix with shape (T, N).
    1.0 means no shock, 0.7 means 30% loss of available labor/capacity.
    """
    N = len(codes)
    mult = np.ones((T, N), dtype=float)

    if shock_code is None or shock_size <= 0:
        return mult

    if shock_code not in codes:
        raise ValueError(f"shock_code '{shock_code}' not found in industry codes.")

    idx = codes.index(shock_code)
    shock_end = min(shock_end, T - 1)
    if shock_start <= shock_end:
        mult[shock_start : shock_end + 1, idx] = 1.0 - shock_size

    return mult

def simulate_linear_stepC(
    baseline,
    T,
    supply_shock_mult,
    mode="linear",
    critical_mask=None,
    important_mask=None,
):
    """
    Step C/D/E simulation:
    - linear
    - pseudo_pbl
    - pbl
    """

    codes = baseline.codes
    N = len(codes)

    A = baseline.A_total
    x0 = baseline.x0_daily
    c0 = baseline.c0_daily
    f0 = baseline.f0_daily
    S = baseline.S0_total.copy()

    if supply_shock_mult is None:
        supply_shock_mult = np.ones((T, N), dtype=float)

    x_hist = np.zeros((T, N), dtype=float)
    d_hist = np.zeros((T, N), dtype=float)
    xcap_hist = np.zeros((T, N), dtype=float)
    xinp_hist = np.zeros((T, N), dtype=float)
    c_hist = np.zeros((T, N), dtype=float)
    f_hist = np.zeros((T, N), dtype=float)

    d_prev = x0.copy()

    for t in range(T):
        # exogenous capacity shock
        x_cap = supply_shock_mult[t] * x0

        # orders
        target_inventory = baseline.S0_total
        O = A * d_prev.reshape(1, -1) + (target_inventory - S) / TAU
        O = np.maximum(O, 0.0)

        # supplier demand
        d = O.sum(axis=1) + c0 + f0

        # input constraint
        if mode == "linear":
            x_inp = compute_input_constraint(
                S=S,
                A=A,
                mode="linear",
            )

        elif mode == "pseudo_pbl":
            x_inp = compute_input_constraint_pseudo_pbl(
                S=S,
                A=A,
                critical_mask=critical_mask,
                important_mask=important_mask,
            )
            
        elif mode == "pbl":
            x_inp = compute_input_constraint_pbl(
                S=S,
                A=A,
                critical_mask=critical_mask,
            )

        else:
            raise ValueError(f"Unknown mode: {mode}")

        # realized output
        x = np.minimum.reduce([x_cap, x_inp, d])

        # proportional rationing
        delivery_ratio = np.ones(N, dtype=float)
        positive_d = d > 1e-15
        delivery_ratio[positive_d] = x[positive_d] / d[positive_d]
        delivery_ratio = np.clip(delivery_ratio, 0.0, 1.0)

        # realized intermediate deliveries
        Z_delivered = O * delivery_ratio.reshape(-1, 1)

        # realized final demand deliveries
        c_realized = c0 * delivery_ratio
        f_realized = f0 * delivery_ratio

        # intended intermediate use
        intended_use = A * x.reshape(1, -1)
        actual_use = np.minimum(intended_use, S)

        # inventory update
        S = np.maximum(S + Z_delivered - actual_use, 0.0)

        # save
        x_hist[t] = x
        d_hist[t] = d
        xcap_hist[t] = x_cap
        xinp_hist[t] = x_inp
        c_hist[t] = c_realized
        f_hist[t] = f_realized

        d_prev = d.copy()

    aggregate_output = x_hist.sum(axis=1)
    aggregate_demand = d_hist.sum(axis=1)
    baseline_aggregate_output = float(x0.sum())
    output_ratio_to_baseline = aggregate_output / baseline_aggregate_output
    sector_output_ratio = safe_divide(x_hist, x0.reshape(1, -1))

    return SimulationResult(
        t=np.arange(T),
        x=x_hist,
        d=d_hist,
        x_cap=xcap_hist,
        x_inp=xinp_hist,
        c_realized=c_hist,
        f_realized=f_hist,
        shock_multiplier=supply_shock_mult,
        aggregate_output=aggregate_output,
        aggregate_demand=aggregate_demand,
        output_ratio_to_baseline=output_ratio_to_baseline,
        sector_output_ratio=sector_output_ratio,
    )

    


# In[11]:


# =========================
# Diagnostics
# =========================

def diagnostic_report_from_baseline(baseline: BEABaseline) -> Dict:
    row_balance = baseline.U_total_annual.sum(axis=1) + baseline.c0_annual + baseline.f0_annual
    commodity_balance_gap = row_balance - baseline.x0_annual

    diag = {
        "n_industries": len(baseline.codes),
        "total_output_annual_sum": float(baseline.x0_annual.sum()),
        "total_labor_comp_annual_sum": float(baseline.l0_annual.sum()),
        "total_pce_annual_sum": float(baseline.c0_annual.sum()),
        "total_other_final_demand_annual_sum": float(baseline.f0_annual.sum()),
        "total_intermediate_annual_sum": float(baseline.U_total_annual.sum()),
        "total_intermediate_import_annual_sum": float(baseline.M_import_annual.sum()),
        "commodity_balance_gap_max_abs": float(np.abs(commodity_balance_gap).max()),
        "propensity_to_consume_m_from_data": float(baseline.c0_annual.sum() / baseline.l0_annual.sum()),
        "mean_import_share_col": float(np.mean(baseline.import_share_col)),
        "max_import_share_col": float(np.max(baseline.import_share_col)),
        "negative_domestic_entries_count_before_clipping": int(baseline.negative_domestic_count),
    }
    return diag


def summarize_steady_state(no_shock_result: SimulationResult) -> Dict:
    ratio = no_shock_result.output_ratio_to_baseline
    return {
        "aggregate_output_ratio_min": float(ratio.min()),
        "aggregate_output_ratio_max": float(ratio.max()),
        "aggregate_output_ratio_last": float(ratio[-1]),
    }


# In[12]:


# =========================
# Export helpers
# =========================

def export_baseline_tables(baseline: BEABaseline) -> None:
    ensure_dir(OUT_DIR)

    meta = pd.DataFrame({
        "code": baseline.codes,
        "name": baseline.names,
        "x0_annual_musd": baseline.x0_annual,
        "l0_annual_musd": baseline.l0_annual,
        "c0_annual_musd": baseline.c0_annual,
        "f0_annual_musd": baseline.f0_annual,
        "inventory_days": baseline.n_days,
        "import_share_col": baseline.import_share_col,
        "n_critical_inputs": baseline.critical_mask.sum(axis=0),
        "n_important_inputs": baseline.important_mask.sum(axis=0),
        "n_noncritical_inputs": baseline.noncritical_mask.sum(axis=0),
    })
    meta.to_csv(OUT_DIR / "baseline_sector_meta.csv", index=False)

    pd.DataFrame(baseline.U_total_annual, index=baseline.codes, columns=baseline.codes).to_csv(
        OUT_DIR / "U_total_annual.csv"
    )
    pd.DataFrame(baseline.M_import_annual, index=baseline.codes, columns=baseline.codes).to_csv(
        OUT_DIR / "M_import_annual.csv"
    )
    pd.DataFrame(baseline.A_total, index=baseline.codes, columns=baseline.codes).to_csv(
        OUT_DIR / "A_total_daily.csv"
    )
    pd.DataFrame(baseline.critical_mask.astype(int), index=baseline.codes, columns=baseline.codes).to_csv(
        OUT_DIR / "critical_mask.csv"
    )
    pd.DataFrame(baseline.important_mask.astype(int), index=baseline.codes, columns=baseline.codes).to_csv(
        OUT_DIR / "important_mask.csv"
    )

    if baseline.negative_domestic_entries:
        pd.DataFrame(baseline.negative_domestic_entries).to_csv(
            OUT_DIR / "negative_domestic_entries.csv", index=False
        )

def export_simulation(result: SimulationResult, baseline: BEABaseline, prefix: str) -> None:
    ensure_dir(OUT_DIR)

    # Aggregate time series
    agg = pd.DataFrame({
        "t": result.t,
        "aggregate_output_musd_per_day": result.aggregate_output,
        "aggregate_demand_musd_per_day": result.aggregate_demand,
        "aggregate_output_ratio_to_baseline": result.output_ratio_to_baseline,
    })
    agg.to_csv(OUT_DIR / f"{prefix}_aggregate_timeseries.csv", index=False)

    # Sector output ratios
    sector_ratio = pd.DataFrame(result.sector_output_ratio, columns=baseline.codes)
    sector_ratio.insert(0, "t", result.t)
    sector_ratio.to_csv(OUT_DIR / f"{prefix}_sector_output_ratio.csv", index=False)

    # Raw output
    sector_output = pd.DataFrame(result.x, columns=baseline.codes)
    sector_output.insert(0, "t", result.t)
    sector_output.to_csv(OUT_DIR / f"{prefix}_sector_output_musd_per_day.csv", index=False)


def make_plots(no_shock: SimulationResult, shock: SimulationResult, baseline: BEABaseline, shock_code: str) -> None:
    ensure_dir(OUT_DIR)

    # Aggregate comparison
    plt.figure(figsize=(10, 6))
    plt.plot(no_shock.t, no_shock.output_ratio_to_baseline, label="No shock")
    plt.plot(shock.t, shock.output_ratio_to_baseline, label=f"Supply shock: {shock_code}")
    plt.axhline(1.0, linestyle="--", linewidth=1)
    plt.xlabel("Day")
    plt.ylabel("Aggregate output / baseline")
    plt.title("Aggregate output ratio to baseline")
    plt.legend()
    plt.tight_layout()
    plt.savefig(OUT_DIR / "aggregate_output_ratio.png", dpi=160)
    plt.show()
    plt.close()

    # Shocked sector path
    if shock_code in baseline.codes:
        idx = baseline.codes.index(shock_code)
        plt.figure(figsize=(10, 6))
        plt.plot(shock.t, shock.sector_output_ratio[:, idx], label=f"{shock_code} output ratio")
        plt.plot(shock.t, shock.shock_multiplier[:, idx], label=f"{shock_code} capacity multiplier")
        plt.axhline(1.0, linestyle="--", linewidth=1)
        plt.xlabel("Day")
        plt.ylabel("Ratio")
        plt.title(f"Shock path and output path: {shock_code} ({baseline.names[idx]})")
        plt.legend()
        plt.tight_layout()
        plt.savefig(OUT_DIR / f"shock_sector_{shock_code}.png", dpi=160)
        plt.show()
        plt.close()

    # Worst-hit sectors at trough
    trough_t = int(np.argmin(shock.output_ratio_to_baseline))
    trough_ratio = shock.sector_output_ratio[trough_t]
    order = np.argsort(trough_ratio)[:15]

    plt.figure(figsize=(10, 7))
    y = np.arange(len(order))
    labels = [f"{baseline.codes[i]}  {baseline.names[i][:40]}" for i in order]
    vals = trough_ratio[order]
    plt.barh(y, vals)
    plt.yticks(y, labels)
    plt.gca().invert_yaxis()
    plt.xlabel("Output / baseline at aggregate trough")
    plt.title("15 lowest sector output ratios at aggregate trough")
    plt.tight_layout()
    plt.savefig(OUT_DIR / "worst_hit_sectors_at_trough.png", dpi=160)
    plt.show()
    plt.close()


# In[13]:


def make_comparison_plots(
    results_by_mode: dict,
    baseline,
    shock_code: str,
    out_dir: Path,
) -> None:
    ensure_dir(out_dir)

    # 1) aggregate comparison
    plt.figure(figsize=(10, 6))
    for mode, result in results_by_mode.items():
        plt.plot(result.t, result.output_ratio_to_baseline, label=mode)
    plt.axhline(1.0, linestyle="--", linewidth=1)
    plt.xlabel("Day")
    plt.ylabel("Aggregate output / baseline")
    plt.title("Aggregate output comparison across production functions")
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_dir / "compare_aggregate_output_ratio.png", dpi=160)
    plt.show()
    plt.close()

    # 2) shocked sector comparison
    if shock_code in baseline.codes:
        idx = baseline.codes.index(shock_code)
        plt.figure(figsize=(10, 6))
        for mode, result in results_by_mode.items():
            plt.plot(result.t, result.sector_output_ratio[:, idx], label=f"{mode}: output")
        plt.plot(
            next(iter(results_by_mode.values())).t,
            next(iter(results_by_mode.values())).shock_multiplier[:, idx],
            linestyle="--",
            label="capacity multiplier",
        )
        plt.axhline(1.0, linestyle="--", linewidth=1)
        plt.xlabel("Day")
        plt.ylabel("Ratio")
        plt.title(f"Shock-sector output comparison: {shock_code} ({baseline.names[idx]})")
        plt.legend()
        plt.tight_layout()
        plt.savefig(out_dir / f"compare_shock_sector_{shock_code}.png", dpi=160)
        plt.show()
        plt.close()


# In[14]:


# =========================
# Main
# =========================

# Step A: Build Baseline
baseline = build_baseline_from_core(IO_CORE_FILE)

diag = diagnostic_report_from_baseline(baseline)
print("\n=== Baseline diagnostics ===")
print(json.dumps(diag, indent=2))

export_baseline_tables(baseline)

# Step B: no-shock simulation
print("\nRunning Step B: no-shock simulation...")
critical_mask, important_mask, noncritical_mask = build_pseudo_pbl_masks(
    baseline.A_total
)

no_shock_mult = build_supply_shock_path(
    codes=baseline.codes,
    T=T_NO_SHOCK,
    shock_code=None,
)
no_shock_result = simulate_linear_stepC(
    baseline=baseline,
    T=T_NO_SHOCK,
    supply_shock_mult=no_shock_mult,
    mode="linear",
    critical_mask=critical_mask,
    important_mask=important_mask,

)
no_shock_summary = summarize_steady_state(no_shock_result)
print("\n=== No-shock summary ===")
print(json.dumps(no_shock_summary, indent=2))
export_simulation(no_shock_result, baseline, prefix="no_shock")

# Step C: single-sector supply shock
print("\nRunning Step C/D/E: supply shock comparison across production modes...")

critical_mask, important_mask, noncritical_mask = build_pseudo_pbl_masks(
    baseline.A_total
)

shock_mult = build_supply_shock_path(
    codes=baseline.codes,
    T=T_SHOCK,
    shock_code=SHOCK_INDUSTRY_CODE,
    shock_start=SHOCK_START,
    shock_end=SHOCK_END,
    shock_size=SHOCK_SIZE,
)

results_by_mode = {}

for mode in COMPARE_MODES:
    result = simulate_linear_stepC(
        baseline=baseline,
        T=T_SHOCK,
        supply_shock_mult=shock_mult,
        mode=mode,
        critical_mask=critical_mask,
        important_mask=important_mask,
    )
    results_by_mode[mode] = result

    export_simulation(result, baseline, prefix=f"{mode}_shock_{SHOCK_INDUSTRY_CODE}")

make_comparison_plots(
    results_by_mode=results_by_mode,
    baseline=baseline,
    shock_code=SHOCK_INDUSTRY_CODE,
    out_dir=OUT_DIR,
)

# Export a compact metadata file for the run
run_meta = {
    "use_file": USE_FILE,
    "import_file": IMPORT_FILE,
    "days_per_year": DAYS_PER_YEAR,
    "inventory_days": INVENTORY_DAYS,
    "tau": TAU,
    "shock_industry_code": SHOCK_INDUSTRY_CODE,
    "shock_industry_name": baseline.names[baseline.codes.index(SHOCK_INDUSTRY_CODE)] if SHOCK_INDUSTRY_CODE in baseline.codes else None,
    "shock_start": SHOCK_START,
    "shock_end": SHOCK_END,
    "shock_size": SHOCK_SIZE,
    "notes": [
        "Comparison across linear / pseudo_pbl / pbl",
        f"compare_modes = {COMPARE_MODES}",
        f"network_mode = {NETWORK_MODE}",
        "Simulation core uses Use table only",
        "Import matrix is retained for diagnostics and future extension only",
        "No endogenous labor adjustment yet",
        "No income-feedback consumption yet",
    ],
}

with open(OUT_DIR / "run_meta.json", "w", encoding="utf-8") as f:
    json.dump(run_meta, f, ensure_ascii=False, indent=2)

print("\nDone.")
print(f"Outputs written to: {OUT_DIR.resolve()}")
print("Main files:")
print(" - baseline_sector_meta.csv")
print(" - Z_domestic_annual.csv")
print(" - M_import_annual.csv")
print(" - A_domestic_daily.csv")
print(" - no_shock_aggregate_timeseries.csv")
print(f" - shock_{SHOCK_INDUSTRY_CODE}_aggregate_timeseries.csv")
print(" - aggregate_output_ratio.png")
print(f" - shock_sector_{SHOCK_INDUSTRY_CODE}.png")
print(" - worst_hit_sectors_at_trough.png")


# In[15]:


gap = baseline.U_total_annual.sum(axis=1) + baseline.c0_annual + baseline.f0_annual - baseline.x0_annual
gap_ratio = gap / baseline.x0_annual
print(np.max(np.abs(gap)))
print(np.max(np.abs(gap_ratio)))


# In[16]:


print(no_shock_summary)


# In[17]:


for mode, result in results_by_mode.items():
    print(
        mode,
        result.output_ratio_to_baseline.min(),
        result.output_ratio_to_baseline[-1]
    )


# In[18]:


from IPython.display import display

# 例: pbl の結果を見る
result = results_by_mode["pbl"]

# aggregate trough
t_trough = int(np.argmin(result.output_ratio_to_baseline))
print("aggregate trough t =", t_trough)

# worst-hit sectors at trough
worst = pd.DataFrame({
    "code": baseline.codes,
    "name": baseline.names,
    "output_ratio_at_trough": result.sector_output_ratio[t_trough],
}).sort_values("output_ratio_at_trough")

display(worst.head(20))


# In[19]:


shock_idx = baseline.codes.index("325")

downstream = pd.DataFrame({
    "code": baseline.codes,
    "name": baseline.names,
    "use_of_shock_output": baseline.A_total[shock_idx, :]
}).sort_values("use_of_shock_output", ascending=False)

display(downstream.head(20))


# In[20]:


critical_exposure = baseline.A_total[baseline.critical_mask].sum(axis=0)
critical_exposure


# In[21]:


network_summary = pd.DataFrame({
    "code": baseline.codes,
    "name": baseline.names,
    "n_critical_inputs_received": baseline.critical_mask.sum(axis=0),
    "n_important_inputs_received": baseline.important_mask.sum(axis=0),
    "n_critical_outputs_supplied": baseline.critical_mask.sum(axis=1),
    "n_important_outputs_supplied": baseline.important_mask.sum(axis=1),
})

display(
    network_summary.sort_values("n_critical_outputs_supplied", ascending=False).head(20)
)
display(
    network_summary.sort_values("n_important_outputs_supplied", ascending=False).head(20)
)

