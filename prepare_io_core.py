#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import pickle
from pathlib import Path
from typing import Dict, List

import numpy as np
from openpyxl import load_workbook

import warnings
warnings.simplefilter('ignore')

USE_FILE = "Use table.xlsx"
IMPORT_FILE = "Import Matrix.xlsx"
OUT_FILE = "io_core.pkl"
DAYS_PER_YEAR = 365.0


def to_float(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---"}:
            return 0.0
        x = x.replace(",", "")
        return float(x)
    return float(x)


def safe_divide(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    out = np.zeros_like(a, dtype=float)
    np.divide(a, b, out=out, where=np.abs(b) > 1e-15)
    return out


def read_use_table(path: str) -> Dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Table"]

    industry_col_start = 3
    industry_col_end = 73
    row_start = 8
    row_end = 78

    codes = [ws.cell(row_start + i, 1).value for i in range(row_end - row_start + 1)]
    names = [ws.cell(row_start + i, 2).value for i in range(row_end - row_start + 1)]
    col_codes = [ws.cell(6, c).value for c in range(industry_col_start, industry_col_end + 1)]
    col_names = [ws.cell(7, c).value for c in range(industry_col_start, industry_col_end + 1)]

    if codes != col_codes:
        raise ValueError("Row codes and column codes do not align for the 71-industry block.")

    U_total = np.array(
        [
            [to_float(ws.cell(r, c).value) for c in range(industry_col_start, industry_col_end + 1)]
            for r in range(row_start, row_end + 1)
        ],
        dtype=float,
    )

    col_pce = 75
    col_total_final_uses = 95
    col_total_commodity_output = 96
    row_compensation = 82
    row_total_industry_output = 86

    c0 = np.array(
        [to_float(ws.cell(r, col_pce).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )

    total_final_uses = np.array(
        [to_float(ws.cell(r, col_total_final_uses).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )

    f0 = total_final_uses - c0
    f0[np.abs(f0) < 1e-12] = 0.0

    x0 = np.array(
        [to_float(ws.cell(r, col_total_commodity_output).value) for r in range(row_start, row_end + 1)],
        dtype=float,
    )

    l0 = np.array(
        [to_float(ws.cell(row_compensation, c).value) for c in range(industry_col_start, industry_col_end + 1)],
        dtype=float,
    )

    x_industry = np.array(
        [to_float(ws.cell(row_total_industry_output, c).value) for c in range(industry_col_start, industry_col_end + 1)],
        dtype=float,
    )

    return {
        "codes": codes,
        "names": names,
        "col_names": col_names,
        "U_total": U_total,
        "c0": c0,
        "f0": f0,
        "x0_commodity": x0,
        "l0": l0,
        "x0_industry": x_industry,
    }


def read_import_matrix(path: str, expected_codes: List[str]) -> np.ndarray:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Table"]

    industry_col_start = 3
    industry_col_end = 73
    row_start = 8
    row_end = 78

    row_codes = [ws.cell(row_start + i, 1).value for i in range(row_end - row_start + 1)]
    col_codes = [ws.cell(6, c).value for c in range(industry_col_start, industry_col_end + 1)]

    if row_codes != expected_codes or col_codes != expected_codes:
        raise ValueError("Import matrix codes do not align with Use table 71-industry block.")

    M_import = np.array(
        [
            [to_float(ws.cell(r, c).value) for c in range(industry_col_start, industry_col_end + 1)]
            for r in range(row_start, row_end + 1)
        ],
        dtype=float,
    )

    return M_import
    


# In[2]:


use = read_use_table(USE_FILE)
M_import = read_import_matrix(IMPORT_FILE, use["codes"])

U_total_annual = use["U_total"]
x0_annual = use["x0_commodity"]
c0_annual = use["c0"]
f0_annual = use["f0"]
l0_annual = use["l0"]

U_total_daily = U_total_annual / DAYS_PER_YEAR
M_import_daily = M_import / DAYS_PER_YEAR
x0_daily = x0_annual / DAYS_PER_YEAR
c0_daily = c0_annual / DAYS_PER_YEAR
f0_daily = f0_annual / DAYS_PER_YEAR
l0_daily = l0_annual / DAYS_PER_YEAR

A_total = safe_divide(U_total_daily, x0_daily.reshape(1, -1))
import_share_matrix = safe_divide(M_import, U_total_annual)
import_share_col = safe_divide(M_import.sum(axis=0), U_total_annual.sum(axis=0))

payload = {
    "codes": use["codes"],
    "names": use["names"],
    "col_names": use["col_names"],
    "U_total_annual": U_total_annual,
    "M_import_annual": M_import,
    "x0_annual": x0_annual,
    "c0_annual": c0_annual,
    "f0_annual": f0_annual,
    "l0_annual": l0_annual,
    "x0_industry": use["x0_industry"],
    "U_total_daily": U_total_daily,
    "M_import_daily": M_import_daily,
    "x0_daily": x0_daily,
    "c0_daily": c0_daily,
    "f0_daily": f0_daily,
    "l0_daily": l0_daily,
    "A_total": A_total,
    "import_share_matrix": import_share_matrix,
    "import_share_col": import_share_col,
    "days_per_year": DAYS_PER_YEAR,
    "source_use_file": USE_FILE,
    "source_import_file": IMPORT_FILE,
}

with open(OUT_FILE, "wb") as f:
    pickle.dump(payload, f)

print(f"Saved: {OUT_FILE}")


# In[ ]:




